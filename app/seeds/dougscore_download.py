"""Imports"""

import os
import re
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# TODO Add some try catches for errors

URL = "https://docs.google.com/spreadsheets/d/1HcFstlJdQMlMEWhbdKXZWdAzR5RFMtj3kywLQcgkGPw/export"
OUTPUT_DIR = os.path.dirname(__file__)
os.makedirs("dougscore_files", exist_ok=True)
EXCEL_FILE = "dougscore_files/dougscore.xlsx"
CSV_FILE = "dougscore_files/dougscore.csv"

excel_filepath = os.path.join(OUTPUT_DIR, EXCEL_FILE)
csv_filepath = os.path.join(OUTPUT_DIR, CSV_FILE)


def download_google_sheet(filepath):
    """Download the Google sheet and save it to a file"""
    response = requests.get(URL, timeout=20, allow_redirects=True)

    if response.status_code == 200:
        with open(filepath, "wb") as f:
            f.write(response.content)
            print(f"File saved to: {filepath}")
    else:
        print(f"Error downloading Google Sheet: {response.status_code}")


def extract_url_from_hyperlink(hyperlink_string):
    """Use regular expression to extract the URL from the hyperlink string"""
    match = re.search(r"\"(https?://[^\"]+)\"", hyperlink_string)
    if match:
        return match.group(1)
    return None


def remove_empty_rows(sheet):
    """Remove fully empty rows in the sheet"""
    for row in reversed(list(sheet.iter_rows(min_row=1))):
        if all(cell.value is None for cell in row):
            sheet.delete_rows(row[0].row)


def analyze_excel(filepath, column_index, rows_to_delete):
    """Cleanup excel sheet and export it with urls"""

    # Load the Excel workbook (wb) and the desired worksheet (ws)
    # We need data_wb to import as "data_only" so we can get the actual data from the cells
    # not the formulas. We need wb so we can get the formulas and parse the embedded url links.
    data_wb = load_workbook(filepath, data_only=True)
    data_ws = data_wb["DougScore"]
    wb = load_workbook(filepath)
    ws = wb["DougScore"]

    # TODO Check and see if there is a more effecient way to accomplish this
    # Remove all of the empty rows from the bottom of the sheet
    remove_empty_rows(ws)
    remove_empty_rows(data_ws)

    # Our sheet has several merged cells at the top of the sheet for formatting, un-merge them.
    for merged_cells in list(ws.merged_cells):
        ws.unmerge_cells(range_string=str(merged_cells))
    for merged_cells in list(data_ws.merged_cells):
        data_ws.unmerge_cells(range_string=str(merged_cells))

    # Our sheet has 3 rows of pinned (frozen) panes at the top for formatting, un-freeze them.
    ws.freeze_panes = None
    data_ws.freeze_panes = None

    # Our sheet has 3 rows of formatted heading at the top of the page, delete them.
    ws.delete_rows(1, amount=rows_to_delete)
    data_ws.delete_rows(1, amount=rows_to_delete)

    # Get the column_letter at the location of our target column index so we can
    # append the same calumn letter in the new sheet.
    column_letter = get_column_letter(column_index)

    # For each row in the new sheet iterate over just the target column index
    for row in ws.iter_rows(min_col=column_index, max_col=column_index):

        row_number = row[0].row
        # Check if the value in the cell is not None, if not, continue.
        if row[0].value is not None:
            cell_value = row[0].value

            if isinstance(cell_value, str):
                # If the value is a string, split the hyperlink and extract the URL
                hyperlink_parts = cell_value.split(",")
                url = extract_url_from_hyperlink(hyperlink_parts[0])
                # Save the extracted string the "data_ws", this is the ws we will export.
                data_ws[f"{column_letter}{row_number}"] = str(url)
            else:
                # If the value is not a string, check if it has the 'hyperlink.target' attribute
                if hasattr(row[0], "hyperlink") and hasattr(row[0].hyperlink, "target"):
                    hyperlink_target = row[0].hyperlink.target
                    # Save the hyperlink_target the "data_ws", this is the ws we will export.
                    data_ws[f"{column_letter}{row_number}"] = str(hyperlink_target)
                else:
                    # 'hyperlink.target' attribute is not present? Append an empty string.
                    data_ws[f"{column_letter}{row_number}"] = str("")
        else:
            # Handle the case where the value is None
            data_ws[f"{column_letter}{row_number}"] = str("")

    # Save the modified data workbook
    data_wb.save(filepath)


def convert_excel_to_csv(filepath):
    """Use pandas library to convert sanitized excel file to csv"""
    # Load excel file into a pandas dataframe, with no index column.
    # We specify "no header" row so the first row is not treaded differently from the others.
    df = pd.read_excel(filepath, index_col=None, engine="openpyxl", header=None)
    # Export dataframe to csv file
    df.to_csv(csv_filepath, index=False, header=None)


download_google_sheet(excel_filepath)
analyze_excel(excel_filepath, 17, 3)
convert_excel_to_csv(excel_filepath)
