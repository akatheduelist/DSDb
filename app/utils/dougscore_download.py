import os
import re
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# TODO Add some try catches for errors

URL = "https://docs.google.com/spreadsheets/d/1HcFstlJdQMlMEWhbdKXZWdAzR5RFMtj3kywLQcgkGPw/export"

OUTPUT_DIR = os.path.dirname(__file__)
EXCEL_FILE = "../seeds/dougscore_files/dougscore.xlsx"
CSV_FILE = "../seeds/dougscore_files/dougscore.csv"

excel_filepath = os.path.join(OUTPUT_DIR, EXCEL_FILE)
csv_filepath = os.path.join(OUTPUT_DIR, CSV_FILE)


def download_google_sheet(filepath):
    """Download the Google sheet and save it to a file"""
    response = requests.get(URL, timeout=20, allow_redirects=True)

    if response.status_code == 200:
        with open(filepath, "wb") as f:
            f.write(response.content)
            print(f"File saved to: {filepath}")
    else:
        print(f"Error downloading Google Sheet: {response.status_code}")


def extract_url_from_hyperlink(hyperlink_string):
    """Use regular expression to extract the URL from the hyperlink string"""
    match = re.search(r"\"(https?://[^\"]+)\"", hyperlink_string)
    if match:
        return match.group(1)
    return None


def remove_empty_rows(sheet):
    """Remove fully empty rows in the sheet"""
    for row in reversed(list(sheet.iter_rows(min_row=1))):
        if all(cell.value is None for cell in row):
            sheet.delete_rows(row[0].row)


def analyze_excel(filepath, column_index, rows_to_delete):
    """Cleanup excel sheet and export it with urls"""
    # Load the Excel workbook (wb) and the desired worksheet (ws)
    wb = load_workbook(filepath)
    ws = wb["DougScore"]

    remove_empty_rows(ws)

    for merged_cells in list(ws.merged_cells):
        ws.unmerge_cells(range_string=str(merged_cells))

    ws.freeze_panes = None
    ws.delete_rows(1, amount=rows_to_delete)
    ws.insert_rows(1)
    new_column_index = column_index + 1
    ws.insert_cols(new_column_index)
    new_column_letter = get_column_letter(new_column_index)

    for row in ws.iter_rows(min_col=column_index, max_col=column_index):
        # Check if the value in the cell is not None
        if row[0].value is not None:
            cell_value = row[0].value
            row_number = row[
                0
            ].row  # Access the row number directly from the row object

            if isinstance(cell_value, str):
                # If the value is a string, split the hyperlink and extract the URL
                hyperlink_parts = cell_value.split(",")
                url = extract_url_from_hyperlink(hyperlink_parts[0])
                ws[f"{new_column_letter}{row_number}"] = str(url)
            else:
                # If the value is not a string, check if it has the 'target' attribute
                if hasattr(row[0], "hyperlink") and hasattr(row[0].hyperlink, "target"):
                    hyperlink_target = row[0].hyperlink.target
                    ws[f"{new_column_letter}{row_number}"] = str(hyperlink_target)
                else:
                    # Handle the case where 'target' attribute is not present
                    ws[f"{new_column_letter}{row_number}"] = str("")
        else:
            # Handle the case where the value is None
            pass

    # Save the modified workbook
    wb.save(filepath)


download_google_sheet(excel_filepath)
analyze_excel(excel_filepath, 17, 3)

col_names = [
    "year",
    "make",
    "model",
    "weekend_styling",
    "weekend_acceleration",
    "weekend_handling",
    "weekend_funfactor",
    "weekend_coolfactor",
    "weekend_total",
    "daily_features",
    "daily_comfort",
    "daily_quality",
    "daily_practical",
    "daily_value",
    "daily_total",
    "dougscore_total",
    "video_runtime",
    "video_link",
    "filming_city",
    "filming_state",
    "vehicle_country",
]

df = pd.read_excel(excel_filepath)
df.columns = col_names
df.to_csv(csv_filepath)
